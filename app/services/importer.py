"""从 gongzi.xlsx 迁移数据到 SQLite，并执行汇总对账。"""
from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..core import repository
from ..core.schema import seed_insurance_items

EXPECTED_ANCHORS = {
    "工资汇总": 362173.758,
    "收入汇总": 407775.778,
    "存款汇总": 96700.0,
    "投资总持仓": 93228.595,
    "总收益率": 0.061708266653594857,
}


class MigrationError(Exception):
    pass


def _num(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rate(text: str) -> float:
    match = re.search(r"\((\d+(?:\.\d+)?)%\)", text)
    return float(match.group(1)) / 100.0 if match else 0.0


def _parse_monthly_sheet(ws) -> list[dict[str, Any]]:
    def header_texts(row_number: int) -> dict[int, str]:
        texts = {}
        for col, cell in enumerate(ws[row_number], start=1):
            text = _text(cell.value)
            if text:
                texts[col] = text
        return texts

    row2 = header_texts(2)
    row3 = header_texts(3)

    def find_row3(label: str) -> int | None:
        for col, text in row3.items():
            if text == label:
                return col
        return None

    salary_col = find_row3("月工资")
    bonus_col = find_row3("年终奖")
    subsidy_col = find_row3("各类补贴")
    reimb_col = find_row3("报销")
    note_col = find_row3("备注")
    rent_col = find_row3("房租")
    utilities_col = find_row3("水电")
    deposit_col = next(
        (col for col, text in row2.items() if text == "强制存款"), None
    )
    deposit_note_col = next(
        (col for col, text in row3.items() if text == "备注" and col > (deposit_col or 0)),
        None,
    )
    housing_note_col = next(
        (col for col, text in row3.items() if text == "备注" and col > (rent_col or 0)),
        None,
    )

    records: list[dict[str, Any]] = []
    for r, row in enumerate(ws.iter_rows(min_row=4, max_row=15), start=4):
        month_val = row[0].value
        if not isinstance(month_val, (int, float)) or int(month_val) not in range(1, 13):
            continue
        month = int(month_val)

        def cell(col: int | None) -> Any:
            if col is None:
                return 0.0
            return ws.cell(row=r, column=col).value

        deposit_raw = _num(cell(deposit_col))
        records.append(
            {
                "month": month,
                "salary": _num(cell(salary_col)),
                "year_end_bonus": _num(cell(bonus_col)),
                "subsidies": _num(cell(subsidy_col)),
                "reimbursements": _num(cell(reimb_col)),
                "income_note": _text(cell(note_col)),
                "rent": _num(cell(rent_col)),
                "utilities": _num(cell(utilities_col)),
                "housing_note": _text(cell(housing_note_col)),
                # Excel 中负数为存入、正数为提取；数据库反过来存正=存入、负=提取。
                "forced_deposit": -deposit_raw,
                "deposit_note": _text(cell(deposit_note_col)),
            }
        )
    return records


def _parse_large_items(ws) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    markers = []
    for r, row in enumerate(ws.iter_rows(), start=1):
        for c, cell in enumerate(row, start=1):
            text = _text(cell.value)
            if text in ("其他大笔消费", "其他大笔收入"):
                markers.append((r, c, text))

    for start_row, start_col, marker in markers:
        item_type = "income" if "收入" in marker else "expense"
        date_col, name_col, amount_col = start_col + 2, start_col + 3, start_col + 4
        for r in range(start_row + 1, ws.max_row + 1):
            name = _text(ws.cell(row=r, column=name_col).value)
            date = _text(ws.cell(row=r, column=date_col).value)
            amount = _num(ws.cell(row=r, column=amount_col).value)
            if name == "总计":
                break
            if not name and not date:
                continue
            items.append(
                {
                    "item_type": item_type,
                    "item_date": date,
                    "name": name,
                    "amount": amount,
                    "note": "",
                }
            )
    return items


def _parse_holdings(ws) -> list[dict[str, Any]]:
    header_rows: list[int] = []
    for r, row in enumerate(ws.iter_rows(), start=1):
        texts = [_text(c.value) for c in row]
        if "投资渠道" in texts and "名称" in texts:
            header_rows.append(r)

    def header_texts(row_number: int) -> dict[int, str]:
        texts = {}
        for col, cell in enumerate(ws[row_number], start=1):
            text = _text(cell.value)
            if text:
                texts[col] = text
        return texts

    holdings: list[dict[str, Any]] = []
    for idx, header_row in enumerate(header_rows):
        end_row = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row
        cells = header_texts(header_row)

        channel_col = next((c for c, t in cells.items() if t == "投资渠道"), None)
        name_col = next((c for c, t in cells.items() if t == "名称"), None)
        holding_col = next((c for c, t in cells.items() if t == "持仓"), None)
        profit_col = next((c for c, t in cells.items() if t == "持有收益"), None)
        cumulative_col = next((c for c, t in cells.items() if t == "累计收益"), None)
        rate_col = next(
            (c for c, t in cells.items() if t in ("持有收益率", "收益率")), None
        )
        plan_col = next((c for c, t in cells.items() if t == "定投"), None)
        time_col = next((c for c, t in cells.items() if t == "定投时间"), None)
        category_col = channel_col - 1 if channel_col else None

        if "股票" in _text(ws.cell(row=header_row, column=category_col or 1).value):
            current_category = "股票"
        else:
            current_category = None
        current_channel = ""

        for r in range(header_row + 1, end_row):
            name = _text(ws.cell(row=r, column=name_col).value)
            if not name or "汇总" in name:
                continue
            cat_raw = (
                _text(ws.cell(row=r, column=category_col).value)
                if category_col
                else ""
            )
            if cat_raw and "投资汇总" not in cat_raw:
                current_category = "股票" if "股票" in cat_raw else cat_raw
            channel = _text(ws.cell(row=r, column=channel_col).value)
            if channel:
                current_channel = channel

            holding = _num(ws.cell(row=r, column=holding_col).value)
            profit = _num(ws.cell(row=r, column=profit_col).value)
            cumulative = (
                _num(ws.cell(row=r, column=cumulative_col).value)
                if cumulative_col
                else profit
            )
            rate_value = (
                ws.cell(row=r, column=rate_col).value if rate_col else None
            )
            cost_value = None
            if "股票" in (current_category or ""):
                cost_col = next(
                    (c for c, t in cells.items() if t == "持仓" and c != holding_col),
                    None,
                )
                if cost_col:
                    cost_value = _num(ws.cell(row=r, column=cost_col).value)

            holdings.append(
                {
                    "category": current_category or "其他",
                    "channel": current_channel,
                    "name": name,
                    "holding_value": holding,
                    "holding_profit": profit,
                    "cumulative_profit": cumulative,
                    "return_rate": float(rate_value) if isinstance(rate_value, (int, float)) else None,
                    "cost_basis": cost_value,
                    "invest_plan": _text(ws.cell(row=r, column=plan_col).value) if plan_col else "",
                    "invest_time": _text(ws.cell(row=r, column=time_col).value) if time_col else "",
                    "note": "",
                }
            )
    return holdings


def _parse_insurance_params(ws, header_row: int) -> dict[str, Any] | None:
    if header_row is None:
        return None
    params: dict[str, Any] = {
        "base": 0.0,
        "monthly_salary": 0.0,
        "thirteenth_month_months": 1.0,
        "year_end_bonus_months": 1.0,
        "thirteenth_amount": 0.0,
        "year_end_bonus_amount": 0.0,
        "housing_subsidy": 0.0,
        "housing_fund_personal_rate": 0.0,
        "housing_fund_company_rate": 0.0,
        "pension_personal_rate": 0.0,
        "pension_company_rate": 0.0,
        "medical_personal_rate": 0.0,
        "medical_company_rate": 0.0,
        "big_medical_personal": 0.0,
        "big_medical_company": 0.0,
        "maternity_personal_rate": 0.0,
        "maternity_company_rate": 0.0,
        "injury_personal_rate": 0.0,
        "injury_company_rate": 0.0,
        "unemployment_personal_rate": 0.0,
        "unemployment_company_rate": 0.0,
    }
    current_category = ""
    for r in range(header_row, min(header_row + 22, ws.max_row + 1)):
        a_text = _text(ws.cell(row=r, column=1).value)
        b_text = _text(ws.cell(row=r, column=2).value)
        c_text = _text(ws.cell(row=r, column=3).value)
        d_value = _num(ws.cell(row=r, column=4).value)
        if not c_text and a_text != "工资/13薪/年终奖":
            continue
        if a_text:
            current_category = a_text
        elif b_text in ("大额医疗费用补助", "生育保险"):
            current_category = b_text
        combined = a_text + b_text + c_text

        if c_text == "五险基数":
            params["base"] = d_value
        elif c_text == "租房补贴":
            params["housing_subsidy"] = d_value
        elif a_text == "工资/13薪/年终奖":
            params["monthly_salary"] = d_value
        elif "大额医疗" in combined or current_category == "大额医疗费用补助":
            if "个人" in c_text:
                params["big_medical_personal"] = d_value
            else:
                params["big_medical_company"] = d_value
        elif "生育" in combined or current_category == "生育保险":
            rate = _rate(c_text)
            if "个人" in c_text:
                params["maternity_personal_rate"] = rate
            else:
                params["maternity_company_rate"] = rate
        elif "工伤" in combined or current_category in ("工伤险", "工伤"):
            rate = _rate(c_text)
            if "个人" in c_text:
                params["injury_personal_rate"] = rate
            else:
                params["injury_company_rate"] = rate
        elif "失业" in combined or current_category in ("失业保险", "失业"):
            rate = _rate(c_text)
            if "个人" in c_text:
                params["unemployment_personal_rate"] = rate
            else:
                params["unemployment_company_rate"] = rate
        elif "公积金" in combined or current_category == "公积金":
            rate = _rate(c_text)
            if "个人" in c_text:
                params["housing_fund_personal_rate"] = rate
            else:
                params["housing_fund_company_rate"] = rate
        elif "养老" in combined or current_category == "养老":
            rate = _rate(c_text)
            if "个人" in c_text:
                params["pension_personal_rate"] = rate
            else:
                params["pension_company_rate"] = rate
        elif "医保" in combined or current_category == "医保":
            rate = _rate(c_text)
            if "个人" in c_text:
                params["medical_personal_rate"] = rate
            else:
                params["medical_company_rate"] = rate
    params["thirteenth_amount"] = params["monthly_salary"]
    params["year_end_bonus_amount"] = params["monthly_salary"]
    return params


def _read_reference_anchors(ws) -> dict[str, float]:
    ref = {
        "工资汇总": _num(ws["D1"].value),
        "收入汇总": _num(ws["D5"].value),
        "存款汇总": _num(ws["D8"].value),
        "投资总持仓": _num(ws["D16"].value),
        "总收益率": _num(ws["G16"].value),
    }
    return ref


def _verify_anchors(conn: sqlite3.Connection, ref: dict[str, float]) -> None:
    wages = conn.execute("SELECT COALESCE(SUM(salary), 0) FROM monthly_records").fetchone()[0]
    income = conn.execute(
        "SELECT COALESCE(SUM(salary + year_end_bonus + subsidies + reimbursements), 0) "
        "FROM monthly_records"
    ).fetchone()[0]
    deposits = conn.execute("SELECT COALESCE(SUM(forced_deposit), 0) FROM monthly_records").fetchone()[0]
    holding = conn.execute("SELECT COALESCE(SUM(holding_value), 0) FROM holdings").fetchone()[0]
    cumulative = conn.execute("SELECT COALESCE(SUM(cumulative_profit), 0) FROM holdings").fetchone()[0]

    computed = {
        "工资汇总": float(wages),
        "收入汇总": float(income),
        "存款汇总": float(deposits),
        "投资总持仓": float(holding),
        "总收益率": float(cumulative) / float(holding) if holding else 0.0,
    }
    errors = []
    for key, expected in ref.items():
        got = computed[key]
        tolerance = 1e-4 if key == "总收益率" else 0.01
        if abs(got - expected) > tolerance:
            errors.append(f"{key}: 软件={got:.6f}, Excel={expected:.6f}")
    if errors:
        raise MigrationError("迁移对账失败：\n" + "\n".join(errors))


def import_xlsx(conn: sqlite3.Connection, xlsx_path: str | Path) -> dict[str, Any]:
    """将 Excel 全量导入数据库；失败时回滚，不生成半成品数据。"""
    path = Path(xlsx_path)
    if not path.exists():
        raise MigrationError(f"找不到文件：{path}")
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        summary_ws = wb["汇总"]
        repository.clear_imported_data(conn)

        year_names = [name for name in wb.sheetnames if name.isdigit()]
        for year_name in sorted(year_names, key=int):
            year = int(year_name)
            year_id = repository.ensure_year(conn, year)
            repository.upsert_monthly_records(
                conn, year_id, _parse_monthly_sheet(wb[year_name])
            )
            for item in _parse_large_items(wb[year_name]):
                repository.add_large_item(conn, year_id, item)

        repository.replace_holdings(conn, _parse_holdings(summary_ws))

        header_row = None
        for row in summary_ws.iter_rows(min_col=1, max_col=1):
            if re.fullmatch(r"收入\(\d{4}\)", _text(row[0].value)):
                header_row = row[0].row
                break
        params = _parse_insurance_params(summary_ws, header_row)
        if params:
            year_match = re.search(r"(\d{4})", _text(summary_ws.cell(row=header_row, column=1).value))
            if year_match:
                year_id = repository.ensure_year(conn, int(year_match.group(1)))
                repository.upsert_insurance_params(conn, year_id, params)
                seed_insurance_items(conn)

        ref = _read_reference_anchors(summary_ws)
        _verify_anchors(conn, ref)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        wb.close()

    return {
        "years": year_names,
        "anchors": EXPECTED_ANCHORS,
        "verified": True,
    }
